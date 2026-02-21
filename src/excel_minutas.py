from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path

import models

HEADERS = [
    "minuta",
    "alimento",
    "gramos_grupo_1",
    "gramos_grupo_2",
]

LOGGER = logging.getLogger(__name__)


@dataclass
class ImportSummary:
    rows_processed: int = 0
    rows_imported: int = 0
    minutas_created: int = 0
    minutas_updated: int = 0
    items_upserted: int = 0
    foods_detected: int = 0
    unknown_food_rows: int = 0
    empty_food_rows: int = 0
    unknown_foods: list[str] | None = None

    def __post_init__(self) -> None:
        if self.unknown_foods is None:
            self.unknown_foods = []


def _normalize_header(value: object) -> str:
    return models.normalize_food_name(str(value or ""))


def _validate_headers(ws: object) -> None:
    header_values = [cell.value for cell in ws[1]]
    normalized_headers = [_normalize_header(value) for value in header_values if value is not None]
    expected = {models.normalize_food_name(header) for header in HEADERS}
    missing = [header for header in HEADERS if models.normalize_food_name(header) not in normalized_headers]
    if missing:
        raise ValueError(
            "La plantilla no contiene todas las columnas requeridas. "
            f"Faltan: {', '.join(missing)}. "
            f"Columnas detectadas: {', '.join(sorted(expected.intersection(normalized_headers)))}"
        )


def export_template(path: str | Path) -> Path:
    try:
        from openpyxl import Workbook
    except ModuleNotFoundError as exc:
        raise RuntimeError("Falta la dependencia 'openpyxl'. Instala requirements.txt") from exc

    output_path = Path(path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Minutas"

    ws.append(HEADERS)

    alimentos = models.list_alimentos()
    alimentos_nombres = sorted([a["nombre"] for a in alimentos], key=lambda x: x.lower())
    for nombre in alimentos_nombres:
        ws.append(["", nombre, "", ""])

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16

    ws2 = wb.create_sheet(title="Instrucciones")
    ws2.append(["Cómo usar la plantilla"])
    ws2.append([
        "1) Completa la columna 'minuta' con el nombre de la minuta (puedes repetirlo en varias filas)."
    ])
    ws2.append([
        "2) El sistema compara alimentos con normalización (espacios/tildes/mayúsculas)."
    ])
    ws2.append([
        "3) Diligencia 'gramos_grupo_1' y 'gramos_grupo_2' con números mayores a 0."
    ])
    ws2.append([
        "4) Puedes añadir más filas para otras minutas."
    ])

    wb.save(output_path)
    return output_path


def import_minutas(path: str | Path) -> ImportSummary:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise RuntimeError("Falta la dependencia 'openpyxl'. Instala requirements.txt") from exc

    wb = load_workbook(filename=Path(path), data_only=True)
    ws = wb["Minutas"] if "Minutas" in wb.sheetnames else wb.active
    _validate_headers(ws)

    summary = ImportSummary()

    alimentos = {models.normalize_food_name(a["nombre"]): a["id"] for a in models.list_alimentos()}
    minutas = {models.normalize_name(m["nombre"]).lower(): m["id"] for m in models.list_minutas()}

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        minuta_raw, alimento_raw, gramos1_raw, gramos2_raw = (row + (None,) * 4)[:4]

        if not any([minuta_raw, alimento_raw, gramos1_raw, gramos2_raw]):
            continue

        summary.rows_processed += 1

        minuta_name = models.normalize_name(str(minuta_raw or ""))
        alimento_name = models.normalize_name(str(alimento_raw or ""))
        if not minuta_name:
            raise ValueError(f"Fila {idx}: 'minuta' es obligatoria.")

        if not alimento_name:
            summary.empty_food_rows += 1
            LOGGER.warning("Fila %s ignorada por alimento vacío.", idx)
            continue

        if gramos1_raw in (None, "") or gramos2_raw in (None, ""):
            continue

        try:
            gramos_1 = float(str(gramos1_raw).replace(",", "."))
            gramos_2 = float(str(gramos2_raw).replace(",", "."))
        except Exception as exc:
            raise ValueError(f"Fila {idx}: gramos inválidos para '{alimento_name}'.") from exc

        if gramos_1 <= 0 or gramos_2 <= 0:
            continue

        alimento_key = models.normalize_food_name(alimento_name)
        alimento_id = alimentos.get(alimento_key)
        if alimento_id is None:
            summary.unknown_food_rows += 1
            if alimento_name not in summary.unknown_foods:
                summary.unknown_foods.append(alimento_name)
            LOGGER.warning("Alimento no encontrado en fila %s: %s", idx, alimento_name)
            continue

        summary.foods_detected += 1

        minuta_key = minuta_name.lower()
        if minuta_key not in minutas:
            minuta_id = models.create_minuta(minuta_name)
            minutas[minuta_key] = minuta_id
            summary.minutas_created += 1
        else:
            minuta_id = minutas[minuta_key]
            summary.minutas_updated += 1

        models.add_or_update_item(minuta_id, alimento_id, gramos_1, gramos_2)
        summary.items_upserted += 1
        summary.rows_imported += 1

    return summary
